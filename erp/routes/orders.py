from flask import (
    Blueprint,
    render_template,
    redirect,
    url_for,
    session,
    request,
    send_file,
    Response,
    stream_with_context,
    current_app,
    jsonify,
)
from flask_wtf import FlaskForm
from wtforms import SelectField, IntegerField, BooleanField, SubmitField, StringField
from wtforms.validators import DataRequired, NumberRange

from sqlalchemy import text
from io import BytesIO, StringIO
import csv
from openpyxl import Workbook
from db import get_db
from erp.utils import login_required, has_permission, idempotency_key_required
from erp.sort_utils import sanitize_sort, sanitize_direction

bp = Blueprint("orders", __name__, url_prefix="/orders")


@bp.route("/put_order", methods=["GET", "POST"])
@login_required
@idempotency_key_required
def put_order():
    if not has_permission("put_order"):
        return redirect(url_for("main.dashboard"))

    class OrderForm(FlaskForm):
        item_id = SelectField("Item", coerce=int, validators=[DataRequired()])
        quantity = IntegerField(
            "Quantity", validators=[DataRequired(), NumberRange(min=1)]
        )
        customer = StringField("Customer", validators=[DataRequired()])
        vat_exempt = BooleanField("VAT Exempt")
        submit = SubmitField("Submit Order")

    conn = get_db()
    cur = conn.execute(text("SELECT id, item_code FROM inventory"))
    form = OrderForm()
    form.item_id.choices = [(row[0], row[1]) for row in cur.fetchall()]
    if form.validate_on_submit():
        item_id = form.item_id.data
        quantity = form.quantity.data
        customer = form.customer.data
        vat_exempt = form.vat_exempt.data
        user = (
            session.get("username")
            if session.get("role") != "Client"
            else session["tin"]
        )
        conn.execute(
            text(
                "INSERT INTO orders (item_id, quantity, customer, sales_rep, vat_exempt, status) "
                "VALUES (:item_id, :quantity, :customer, :sales_rep, :vat_exempt, 'pending')"
            ),
            {
                "item_id": item_id,
                "quantity": quantity,
                "customer": customer,
                "sales_rep": user,
                "vat_exempt": vat_exempt,
            },
        )
        conn.commit()
        conn.close()
        return redirect(url_for("main.dashboard"))
    conn.close()
    return render_template("put_order.html", form=form)


@bp.route("/orders")
@login_required
def orders():
    if not has_permission("view_orders"):
        return redirect(url_for("main.dashboard"))
    conn = get_db()
    pending_orders = conn.execute(
        text("SELECT * FROM orders WHERE status = 'pending' ORDER BY id DESC LIMIT 5")
    ).fetchall()
    approved_orders = conn.execute(
        text("SELECT * FROM orders WHERE status = 'approved' ORDER BY id DESC LIMIT 5")
    ).fetchall()
    rejected_orders = conn.execute(
        text("SELECT * FROM orders WHERE status = 'rejected' ORDER BY id DESC LIMIT 5")
    ).fetchall()
    conn.close()
    return render_template(
        "orders.html",
        pending_orders=pending_orders,
        approved_orders=approved_orders,
        rejected_orders=rejected_orders,
    )


ALLOWED_SORTS = {"id", "item_id", "quantity", "customer", "status"}


def _build_query(sort: str, direction: str, limit: int | None = None, offset: int | None = None):
    sort = sanitize_sort(sort, ALLOWED_SORTS)
    direction = sanitize_direction(direction)
    order_sql = "DESC" if direction == "desc" else "ASC"
    sql = f"SELECT id, item_id, quantity, customer, status FROM orders ORDER BY {sort} {order_sql}"
    if limit is not None:
        sql += " LIMIT :limit"
    if offset is not None:
        sql += " OFFSET :offset"
    return text(sql)


def _export_rows(rows, fmt):
    headers = ["id", "item_id", "quantity", "customer", "status"]
    if fmt == "xlsx":
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        for r in rows:
            ws.append([r[h] for h in headers])
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        return send_file(
            bio,
            as_attachment=True,
            download_name="orders.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def generate():
        sio = StringIO()
        writer = csv.writer(sio)
        writer.writerow(headers)
        yield sio.getvalue()
        sio.seek(0)
        sio.truncate(0)
        for r in rows:
            writer.writerow([r[h] for h in headers])
            yield sio.getvalue()
            sio.seek(0)
            sio.truncate(0)

    return Response(
        stream_with_context(generate()),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=orders.csv"},
    )


def _iter_rows(conn, stmt):
    cur = conn.execute(stmt)
    columns = [c[0] for c in cur.description]
    try:
        for row in cur:
            yield dict(zip(columns, row))
    finally:
        cur.close()
        conn.close()


@bp.route("/list")
@login_required
def orders_list():
    if not has_permission("view_orders"):
        return redirect(url_for("main.dashboard"))
    sort = sanitize_sort(request.args.get("sort", "id"), ALLOWED_SORTS)
    direction = sanitize_direction(request.args.get("dir", "asc"))
    limit = min(int(request.args.get("limit", 20)), 100)
    offset = int(request.args.get("offset", 0))
    conn = get_db()
    cur = conn.execute(
        _build_query(sort, direction, limit, offset),
        {"limit": limit, "offset": offset},
    )
    columns = [c[0] for c in cur.description]
    rows = [dict(zip(columns, r)) for r in cur.fetchall()]
    cur.close()
    conn.close()
    if current_app.config.get("TESTING"):
        return jsonify(rows)
    next_offset = offset + limit if len(rows) == limit else None
    prev_offset = offset - limit if offset - limit >= 0 else None
    return render_template(
        "orders_list.html",
        orders=rows,
        sort=sort,
        direction=direction,
        limit=limit,
        offset=offset,
        next_offset=next_offset,
        prev_offset=prev_offset,
    )


@bp.route("/export.csv")
@login_required
def export_orders_csv():
    if not has_permission("view_orders"):
        return redirect(url_for("main.dashboard"))
    sort = sanitize_sort(request.args.get("sort", "id"), ALLOWED_SORTS)
    direction = sanitize_direction(request.args.get("dir", "asc"))
    conn = get_db()
    rows = _iter_rows(conn, _build_query(sort, direction))
    return _export_rows(rows, "csv")


@bp.route("/export.xlsx")
@login_required
def export_orders_xlsx():
    if not has_permission("view_orders"):
        return redirect(url_for("main.dashboard"))
    sort = sanitize_sort(request.args.get("sort", "id"), ALLOWED_SORTS)
    direction = sanitize_direction(request.args.get("dir", "asc"))
    conn = get_db()
    rows = _iter_rows(conn, _build_query(sort, direction))
    return _export_rows(rows, "xlsx")
