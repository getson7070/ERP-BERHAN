from flask import (
    Blueprint,
    render_template,
    redirect,
    url_for,
    session,
    flash,
    request,
    send_file,
    Response,
    stream_with_context,
    current_app,
    jsonify,
)
from flask_wtf import FlaskForm
from wtforms import StringField, SelectField, DateField, SubmitField
from wtforms.validators import DataRequired

from sqlalchemy import text
from db import get_db
from erp.utils import login_required, has_permission
from erp.utils import sanitize_sort, sanitize_direction
from io import BytesIO, StringIO
import csv
from openpyxl import Workbook

bp = Blueprint("tenders", __name__, url_prefix="/tenders")

ALLOWED_SORTS = {
    "id",
    "type_name",
    "description",
    "due_date",
    "workflow_state",
    "result",
    "awarded_to",
    "award_date",
    "username",
    "institution",
    "envelope_type",
}


def _build_query(sort: str, direction: str, limit: int | None = None, offset: int | None = None):
    columns = {
        "id": "t.id",
        "type_name": "tt.type_name",
        "description": "t.description",
        "due_date": "t.due_date",
        "workflow_state": "t.workflow_state",
        "result": "t.result",
        "awarded_to": "t.awarded_to",
        "award_date": "t.award_date",
        "username": "t.username",
        "institution": "t.institution",
        "envelope_type": "t.envelope_type",
    }
    sort_col = columns[sanitize_sort(sort, ALLOWED_SORTS, "due_date")]
    order_sql = "DESC" if sanitize_direction(direction) == "desc" else "ASC"
    sql = (
        "SELECT t.id, tt.type_name, t.description, t.due_date, t.workflow_state, t.result, "
        "t.awarded_to, t.award_date, t.username, t.institution, t.envelope_type "
        "FROM tenders t JOIN tender_types tt ON t.tender_type_id = tt.id "
        f"ORDER BY {sort_col} {order_sql}"
    )
    if limit is not None:
        sql += " LIMIT :limit"
    if offset is not None:
        sql += " OFFSET :offset"
    return text(sql)


def _iter_rows(conn, stmt, params=None):
    cur = conn.execute(stmt, params or {})
    columns = [c[0] for c in cur.description]
    try:
        for row in cur:
            yield dict(zip(columns, row))
    finally:
        cur.close()
        conn.close()


def _export_rows(rows, fmt: str):
    headers = [
        "id",
        "type_name",
        "description",
        "due_date",
        "workflow_state",
        "result",
        "awarded_to",
        "award_date",
        "username",
        "institution",
        "envelope_type",
    ]
    if fmt == "xlsx":
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        for r in rows:
            ws.append([r[h] for h in headers])
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        return send_file(
            bio,
            as_attachment=True,
            download_name="tenders.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def generate():
        sio = StringIO()
        writer = csv.writer(sio)
        writer.writerow(headers)
        yield sio.getvalue()
        sio.seek(0)
        sio.truncate(0)
        for r in rows:
            writer.writerow([r[h] for h in headers])
            yield sio.getvalue()
            sio.seek(0)
            sio.truncate(0)

    return Response(
        stream_with_context(generate()),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=tenders.csv"},
    )

# ordered states reflecting the full tender lifecycle
WORKFLOW_STATES = [
    "advert_registered",
    "decided_to_register",
    "documents_secured",
    "preparing_documentation",
    "documentation_prepared",
    "document_submitted",
    "opening_minute",
    "evaluated",
    "awarded",
]


@bp.route("/add_tender", methods=["GET", "POST"])
@login_required
def add_tender():
    if not has_permission("add_tender"):
        return redirect(url_for("main.dashboard"))

    class TenderForm(FlaskForm):
        tender_type_id = SelectField(
            "Tender Type", coerce=int, validators=[DataRequired()]
        )
        description = StringField("Description", validators=[DataRequired()])
        due_date = DateField("Due Date", validators=[DataRequired()])
        institution = StringField("Institution")
        envelope_type = SelectField(
            "Envelope Type",
            choices=[
                ("One Envelope", "One Envelope"),
                ("Two Envelope", "Two Envelope"),
            ],
            validators=[DataRequired()],
        )
        private_key = StringField("Private Key")
        tech_key = StringField("Technical Key")
        fin_key = StringField("Financial Key")
        submit = SubmitField("Submit")

    conn = get_db()
    form = TenderForm()
    types = (
        conn.execute(text("SELECT id, type_name FROM tender_types"))
        .mappings()
        .fetchall()
    )
    form.tender_type_id.choices = [(t["id"], t["type_name"]) for t in types]
    if form.validate_on_submit():
        tender_type_id = form.tender_type_id.data
        description = form.description.data
        due_date = form.due_date.data
        institution = form.institution.data
        envelope_type = form.envelope_type.data
        private_key = form.private_key.data if envelope_type == "One Envelope" else None
        tech_key = form.tech_key.data if envelope_type == "Two Envelope" else None
        fin_key = form.fin_key.data if envelope_type == "Two Envelope" else None
        user = (
            session.get("username")
            if session.get("role") != "Client"
            else session["tin"]
        )
        conn.execute(
            text(
                "INSERT INTO tenders (tender_type_id, description, due_date, workflow_state, username, institution, envelope_type, private_key, tech_key, fin_key) "
                "VALUES (:tender_type_id, :description, :due_date, :state, :username, :institution, :envelope_type, :private_key, :tech_key, :fin_key)"
            ),
            {
                "tender_type_id": tender_type_id,
                "description": description,
                "due_date": due_date,
                "state": "advert_registered",
                "username": user,
                "institution": institution,
                "envelope_type": envelope_type,
                "private_key": private_key,
                "tech_key": tech_key,
                "fin_key": fin_key,
            },
        )
        conn.commit()
        conn.close()
        return redirect(url_for("main.dashboard"))
    conn.close()
    return render_template("add_tender.html", form=form)


@bp.route("/list")
@login_required
def tenders_list():
    if not has_permission("tenders_list"):
        return redirect(url_for("main.dashboard"))
    sort = sanitize_sort(request.args.get("sort", "due_date"), ALLOWED_SORTS, "due_date")
    direction = sanitize_direction(request.args.get("dir", "asc"))
    limit = min(int(request.args.get("limit", 20)), 100)
    offset = int(request.args.get("offset", 0))
    conn = get_db()
    stmt = _build_query(sort, direction, limit, offset)
    rows = list(_iter_rows(conn, stmt, {"limit": limit, "offset": offset}))
    next_offset = offset + limit if len(rows) == limit else None
    prev_offset = offset - limit if offset - limit >= 0 else None
    if current_app.config.get("TESTING"):
        return jsonify(rows)
    return render_template(
        "tenders_list.html",
        tenders=rows,
        states=WORKFLOW_STATES,
        sort=sort,
        direction=direction,
        limit=limit,
        offset=offset,
        next_offset=next_offset,
        prev_offset=prev_offset,
    )


@bp.route("/report")
@login_required
def tenders_report():
    if not has_permission("tenders_report"):
        return redirect(url_for("main.dashboard"))
    conn = get_db()
    tenders = (
        conn.execute(
            text(
                "SELECT t.id, tt.type_name, t.description, t.due_date, t.workflow_state, t.result, t.awarded_to, t.award_date, t.username, t.institution, t.envelope_type FROM tenders t JOIN tender_types tt ON t.tender_type_id = tt.id ORDER BY t.due_date ASC"
            )
        )
        .mappings()
        .fetchall()
    )
    conn.close()
    return render_template(
        "tenders_report.html", tenders=tenders, states=WORKFLOW_STATES
    )

@bp.route("/export.csv")
@login_required
def export_tenders_csv():
    if not has_permission("tenders_list"):
        return redirect(url_for("main.dashboard"))
    sort = sanitize_sort(request.args.get("sort", "due_date"), ALLOWED_SORTS, "due_date")
    direction = sanitize_direction(request.args.get("dir", "asc"))
    conn = get_db()
    rows = _iter_rows(conn, _build_query(sort, direction))
    return _export_rows(rows, "csv")


@bp.route("/export.xlsx")
@login_required
def export_tenders_xlsx():
    if not has_permission("tenders_list"):
        return redirect(url_for("main.dashboard"))
    sort = sanitize_sort(request.args.get("sort", "due_date"), ALLOWED_SORTS, "due_date")
    direction = sanitize_direction(request.args.get("dir", "asc"))
    conn = get_db()
    rows = _iter_rows(conn, _build_query(sort, direction))
    return _export_rows(rows, "xlsx")


@bp.route("/<int:tender_id>/advance", methods=["POST"])
@login_required
def advance_tender(tender_id):
    if not has_permission("tenders_list"):
        return redirect(url_for("main.dashboard"))

    conn = get_db()
    tender = conn.execute(
        text("SELECT workflow_state FROM tenders WHERE id = :id"),
        {"id": tender_id},
    ).fetchone()
    if not tender:
        conn.close()
        flash("Tender not found.", "danger")
        return redirect(url_for("tenders.tenders_list"))

    # Support both tuple and mapping rows for different DB drivers
    current = (
        tender[0] if isinstance(tender, tuple) else tender._mapping["workflow_state"]
    )
    idx = WORKFLOW_STATES.index(current)

    if (
        current not in ["opening_minute", "evaluated", "awarded"]
        and idx < len(WORKFLOW_STATES) - 1
    ):
        next_state = WORKFLOW_STATES[idx + 1]
        conn.execute(
            text("UPDATE tenders SET workflow_state = :state WHERE id = :id"),
            {"state": next_state, "id": tender_id},
        )
        flash(f"Tender advanced to {next_state.replace('_', ' ').title()}.", "info")
    elif current == "opening_minute":
        if request.form.get("evaluation_complete"):
            conn.execute(
                text("UPDATE tenders SET workflow_state = :state WHERE id = :id"),
                {"state": "evaluated", "id": tender_id},
            )
            flash("Tender marked as Evaluated.", "info")
        else:
            flash("Evaluation not complete.", "danger")
    elif current == "evaluated":
        result = request.form.get("result")
        awarded_to = request.form.get("awarded_to")
        award_date = request.form.get("award_date")
        if result == "won" and awarded_to and award_date:
            conn.execute(
                text(
                    "UPDATE tenders SET workflow_state = :state, result = :result, awarded_to = :awarded_to, award_date = :award_date WHERE id = :id"
                ),
                {
                    "state": "awarded",
                    "result": result,
                    "awarded_to": awarded_to,
                    "award_date": award_date,
                    "id": tender_id,
                },
            )
            flash("Tender awarded.", "success")
        elif result and result in ["defeat", "rejected", "cancelled"]:
            conn.execute(
                text("UPDATE tenders SET result = :result WHERE id = :id"),
                {"result": result, "id": tender_id},
            )
            flash("Tender result recorded.", "warning")
        else:
            flash("Result and award details required.", "danger")
    else:
        flash("Tender already awarded.", "info")

    conn.commit()
    conn.close()
    return redirect(url_for("tenders.tenders_list"))
